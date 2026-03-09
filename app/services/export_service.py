import io
from datetime import datetime
from pathlib import Path
from ..config import to_jst_str

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..models.tables import SearchJob, RankingResult, ExportJob


def build_ranked_df(db: Session, job_id: int) -> pd.DataFrame:
    rankings = (
        db.query(RankingResult)
        .filter(RankingResult.job_id == job_id)
        .order_by(RankingResult.final_rank)
        .all()
    )
    rows = []
    for r in rankings:
        candidate = r.candidate
        paper = candidate.paper if candidate else None
        score = candidate.score if candidate else None
        if not paper:
            continue
        doi = paper.doi or ""
        doi_url = f"https://doi.org/{doi}" if doi else ""
        rows.append({
            "Rank": r.final_rank,
            "Decision": r.decision,
            "Title": paper.title,
            "Year": paper.year,
            "Journal": paper.journal or "",
            "DOI": doi,
            "DOI Link": doi_url,
            "Total Score": score.total_score if score else 0,
            "Theme Score": score.theme_score if score else 0,
            "Method Score": score.method_score if score else 0,
            "Recency Score": score.recency_score if score else 0,
            "Impact Score": score.impact_score if score else 0,
            "Readability Score": score.readability_score if score else 0,
            "Role Bonus": score.role_bonus if score else 0,
            "Why Relevant": r.reason_text or "",
            "Discovery Path": candidate.discovery_path or "",
            "Citation Count": paper.citation_count,
        })
    return pd.DataFrame(rows)


def export_to_xlsx(db: Session, job: SearchJob, output_dir: str | None = None) -> bytes:
    df = build_ranked_df(db, job.job_id)
    query = job.query

    wb = Workbook()

    # ---- Sheet 1: Summary ----
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary(ws_summary, job, query, df)

    # ---- Sheet 2: Ranked Papers ----
    ws_ranked = wb.create_sheet("Ranked Papers")
    ranked_cols = [
        "Rank", "Decision", "Title", "Year", "Journal", "DOI",
        "Total Score", "Why Relevant", "Discovery Path", "Citation Count"
    ]
    _write_dataframe(ws_ranked, df[ranked_cols] if not df.empty else pd.DataFrame(columns=ranked_cols))

    # ---- Sheet 3: Score Breakdown ----
    ws_score = wb.create_sheet("Score Breakdown")
    score_cols = [
        "Rank", "Title", "Total Score", "Theme Score", "Method Score",
        "Recency Score", "Impact Score", "Readability Score", "Role Bonus"
    ]
    _write_dataframe(ws_score, df[score_cols] if not df.empty else pd.DataFrame(columns=score_cols))

    # ---- Sheet 4: Search Condition ----
    ws_cond = wb.create_sheet("Search Condition")
    _write_search_condition(ws_cond, job, query)

    # バイトとして返す
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    xlsx_bytes = buf.read()

    # ファイルとして保存する場合
    if output_dir:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"ranked_papers_{ts}.xlsx"
        out_path = Path(output_dir) / filename
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with open(out_path, "wb") as f:
            f.write(xlsx_bytes)
        _save_export_log(db, job.job_id, "xlsx", str(out_path))

    return xlsx_bytes


def export_to_csv(db: Session, job_id: int) -> bytes:
    df = build_ranked_df(db, job_id)
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def _write_summary(ws, job, query, df: pd.DataFrame):
    header_font = Font(bold=True, size=12)
    ws["A1"] = "論文探索結果サマリー"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    rows = [
        ("検索タイトル", job.title),
        ("テーマ文", query.theme_text if query else ""),
        ("作成日", to_jst_str(job.created_at)),
        ("候補数", len(df)),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = header_font
        ws.cell(row=i, column=2, value=value)

    ws.cell(row=8, column=1, value="上位5件").font = header_font
    top5 = df.head(5) if not df.empty else pd.DataFrame()
    for j, row in enumerate(top5.itertuples(), start=9):
        ws.cell(row=j, column=1, value=getattr(row, "Rank", ""))
        ws.cell(row=j, column=2, value=getattr(row, "Title", ""))
        ws.cell(row=j, column=3, value=getattr(row, "Total_Score", getattr(row, "Total Score", "")))

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 15


def _write_dataframe(ws, df: pd.DataFrame):
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    doi_col_idx = next(
        (i + 1 for i, c in enumerate(df.columns) if c == "DOI"), None
    )
    link_font = Font(color="0563C1", underline="single")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            # DOI列にハイパーリンクを付与
            if col_idx == doi_col_idx and value:
                cell.hyperlink = f"https://doi.org/{value}"
                cell.font = link_font

    # 列幅調整
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        if col_name in ("Title", "Why Relevant"):
            ws.column_dimensions[letter].width = 50
        elif col_name in ("DOI", "Journal"):
            ws.column_dimensions[letter].width = 30
        else:
            ws.column_dimensions[letter].width = 15

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_search_condition(ws, job, query):
    header_font = Font(bold=True)
    rows = [
        ("Theme Text", query.theme_text if query else ""),
        ("Year From", query.year_from if query else ""),
        ("Year To", query.year_to if query else ""),
        ("Include Terms", query.include_terms if query else ""),
        ("Exclude Terms", query.exclude_terms if query else ""),
        ("Subject Tags", query.subject_tags if query else ""),
        ("Max Candidates", query.max_candidates if query else ""),
    ]
    for i, (label, value) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=label).font = header_font
        ws.cell(row=i, column=2, value=value)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60

    seed_row = len(rows) + 2
    ws.cell(row=seed_row, column=1, value="Seed Papers").font = header_font
    for j, seed in enumerate(job.seed_papers, start=seed_row + 1):
        ws.cell(row=j, column=2, value=seed.doi or seed.url or "")


def _save_export_log(db: Session, job_id: int, fmt: str, path: str):
    log = ExportJob(job_id=job_id, format=fmt, output_path=path)
    db.add(log)
    db.flush()
